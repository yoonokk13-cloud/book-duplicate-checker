"""
도서관 장서 수서 중복 체크 도우미
- 실행: streamlit run book_duplicate_checker.py
- 필요 패키지: pip install streamlit pandas openpyxl requests
"""

import io
import requests
import streamlit as st
import pandas as pd
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ────────────────────────────────────────────
# 페이지 설정
# ────────────────────────────────────────────
st.set_page_config(page_title="도서관 수서 중복 체크", page_icon="📚", layout="wide")
st.title("📚 도서관 장서 수서 중복 체크 도우미")
st.write("알라딘 구입 목록과 도서관 소장 목록을 대조합니다.")

# ────────────────────────────────────────────
# 컬럼 규격 (표준 컬럼명)
# ────────────────────────────────────────────
ALADDIN_REQUIRED = ["서명", "저자", "출판사", "정가", "판매가", "ISBN13", "출간일", "분야"]
LIBRARY_REQUIRED = ["서명", "저자", "출판사", "정가"]

# 키워드 매핑 테이블: 표준 컬럼명 → 인식할 키워드 목록
COLUMN_KEYWORD_MAP = {
    "서명":   ["서명", "도서명", "책명", "제목", "title", "book"],
    "저자":   ["저자", "지은이", "작가", "글쓴이", "author", "writer"],
    "출판사": ["출판사", "출판", "publisher", "발행처", "발행"],
    "정가":   ["정가", "가격", "price"],
    "판매가": ["판매가", "판매", "sale", "할인"],
    "ISBN13": ["isbn13", "isbn", "바코드", "barcode"],
    "출간일": ["출간일", "출판일", "발행일", "date"],
    "분야":   ["분야", "주제", "카테고리", "장르", "category", "subject"],
}

# ────────────────────────────────────────────
# 유틸 함수
# ────────────────────────────────────────────
def clean_isbn(val) -> str:
    return str(val).strip().replace("-", "").replace(" ", "").removesuffix(".0")

def clean_text(val) -> str:
    return (
        str(val).strip().lower()
        .replace(" ", "").replace("-", "").replace(".", "")
        .replace(",", "").replace("·", "").replace("(", "").replace(")", "")
    )

def load_file(uploaded) -> pd.DataFrame | None:
    try:
        if uploaded.name.lower().endswith(".csv"):
            for enc in ["utf-8-sig", "cp949", "euc-kr", "utf-8"]:
                try:
                    uploaded.seek(0)
                    return pd.read_csv(uploaded, encoding=enc, dtype=str)
                except UnicodeDecodeError:
                    continue
            st.error("CSV 인코딩을 인식할 수 없습니다.")
            return None
        else:
            uploaded.seek(0)
            df = pd.read_excel(uploaded, dtype=str)
            return df.dropna(how="all").reset_index(drop=True)
    except Exception as e:
        st.error(f"파일 로드 오류: {e}")
        return None

def to_excel_bytes(df: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="결과")
        ws = writer.sheets["결과"]
        hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        hdr_font = Font(bold=True, color="FFFFFF", size=10)
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws.row_dimensions[1].height = 20
        dup_col_idx = next((i for i, c in enumerate(ws[1], 1) if c.value == "중복여부"), None)
        yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            is_dup = dup_col_idx and "소장 중" in str(row[dup_col_idx - 1].value)
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="center" if cell.column == dup_col_idx else "left",
                    vertical="center",
                )
                if is_dup:
                    cell.fill = yellow
        for col in ws.columns:
            ltr = get_column_letter(col[0].column)
            w = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[ltr].width = max(10, min(w + 3, 50))
        ws.freeze_panes = "A2"
    return output.getvalue()

# ────────────────────────────────────────────
# 기능 1: 컬럼명 자동 매핑
# ────────────────────────────────────────────
def auto_map_columns(df: pd.DataFrame, required: list[str]) -> dict[str, str | None]:
    """
    업로드 파일의 컬럼명을 표준 컬럼명으로 자동 매핑.
    반환: { 표준컬럼명: 실제컬럼명 or None }
    """
    mapping: dict[str, str | None] = {}
    used_cols: set[str] = set()

    for std_col in required:
        keywords = COLUMN_KEYWORD_MAP.get(std_col, [std_col])
        matched = None
        for col in df.columns:
            if col in used_cols:
                continue
            col_norm = col.strip().lower().replace(" ", "")
            for kw in keywords:
                if kw.lower() in col_norm or col_norm in kw.lower():
                    matched = col
                    break
            if matched:
                break
        mapping[std_col] = matched
        if matched:
            used_cols.add(matched)

    return mapping

def apply_mapping(df: pd.DataFrame, mapping: dict[str, str | None]) -> pd.DataFrame:
    """매핑 테이블에 따라 컬럼명을 표준명으로 변환한 DataFrame 반환."""
    rename = {v: k for k, v in mapping.items() if v is not None}
    df = df.rename(columns=rename)
    # 매핑 안 된 표준 컬럼은 빈 컬럼으로 추가
    for std_col, original in mapping.items():
        if original is None and std_col not in df.columns:
            df[std_col] = ""
    return df

def show_mapping_editor(df: pd.DataFrame, required: list[str], label: str) -> pd.DataFrame | None:
    """
    자동 매핑 결과를 보여주고 사용자가 수정할 수 있는 UI.
    확인 완료 시 표준 컬럼명으로 변환된 DataFrame 반환.
    """
    mapping = auto_map_columns(df, required)
    file_cols = ["(없음)"] + df.columns.tolist()

    st.markdown(f"**{label} — 컬럼 매핑 확인**")
    st.caption("자동으로 감지한 결과입니다. 잘못된 항목은 직접 수정하세요.")

    cols = st.columns(len(required))
    final_mapping: dict[str, str | None] = {}

    for i, std_col in enumerate(required):
        guessed = mapping[std_col]
        default_idx = file_cols.index(guessed) if guessed in file_cols else 0
        selected = cols[i].selectbox(
            std_col,
            options=file_cols,
            index=default_idx,
            key=f"map_{label}_{std_col}",
        )
        final_mapping[std_col] = None if selected == "(없음)" else selected

    # 미매핑 필수 컬럼 경고
    missing = [k for k, v in final_mapping.items() if v is None]
    if missing:
        st.warning(f"매핑되지 않은 컬럼: `{'`, `'.join(missing)}` — 해당 항목은 빈 값으로 처리됩니다.")

    return apply_mapping(df.copy(), final_mapping)

# ────────────────────────────────────────────
# 기능 2: 알라딘 API로 ISBN13 자동 조회
# ────────────────────────────────────────────

@st.cache_data(ttl=3600, show_spinner=False)
def fetch_isbn_from_aladin(title: str, author: str, ttb_key: str) -> str | None:
    """
    알라딘 API로 서명+저자 검색 → ISBN13 반환.
    결과는 1시간 캐싱 — 같은 서명은 재실행해도 API 재호출 없음.
    """
    query = f"{title} {author}".strip()
    url = "https://www.aladin.co.kr/ttb/api/ItemSearch.aspx"
    params = {
        "TTBKey": ttb_key,
        "Query": query,
        "QueryType": "Title",
        "MaxResults": 3,
        "start": 1,
        "SearchTarget": "Book",
        "output": "js",
        "Version": "20131101",
    }
    try:
        resp = requests.get(url, params=params, timeout=5)
        data = resp.json()
        items = data.get("item", [])
        if not items:
            return None
        title_norm = clean_text(title)
        for item in items:
            if clean_text(item.get("title", "")) == title_norm:
                return str(item.get("isbn13", "")) or None
        return str(items[0].get("isbn13", "")) or None
    except Exception:
        return None

def enrich_library_with_isbn(df: pd.DataFrame, ttb_key: str) -> pd.DataFrame:
    """
    소장 목록 ISBN13 자동 조회.
    - 병렬 호출(최대 10개 동시)로 속도 향상
    - @st.cache_data로 동일 서명 재조회 방지
    """
    if "ISBN13" not in df.columns:
        df["ISBN13"] = ""

    # 조회가 필요한 행만 추출
    needs_lookup = [
        idx for idx in df.index
        if not (clean_isbn(df.at[idx, "ISBN13"]) not in ("", "nan")
                and len(clean_isbn(df.at[idx, "ISBN13"])) == 13)
    ]

    total   = len(df)
    to_fetch = len(needs_lookup)
    done_count = total - to_fetch

    progress = st.progress(done_count / total, text=f"ISBN 조회 중... (0/{to_fetch})")
    completed = 0

    # 병렬 호출: 최대 10개 동시
    with ThreadPoolExecutor(max_workers=10) as executor:
        future_to_idx = {
            executor.submit(
                fetch_isbn_from_aladin,
                str(df.at[idx, "서명"]),
                str(df.at[idx, "저자"]) if "저자" in df.columns else "",
                ttb_key,
            ): idx
            for idx in needs_lookup
        }

        for future in as_completed(future_to_idx):
            idx = future_to_idx[future]
            isbn = future.result()
            df.at[idx, "ISBN13"] = isbn if isbn else ""

            completed += 1
            progress.progress(
                (done_count + completed) / total,
                text=f"ISBN 조회 중... ({completed}/{to_fetch})",
            )

    progress.empty()
    found = df["ISBN13"].apply(
        lambda x: clean_isbn(x) not in ("", "nan") and len(clean_isbn(x)) == 13
    ).sum()
    st.success(f"ISBN 조회 완료: 전체 {total}건 중 **{found}건** 확인")
    return df

# ────────────────────────────────────────────
# 대조 로직
# ────────────────────────────────────────────
def run_duplicate_check(df_planned: pd.DataFrame, df_library: pd.DataFrame) -> pd.DataFrame:
    """
    우선순위:
    1. 양쪽 모두 ISBN 있으면 → ISBN 대조
    2. 소장 목록 ISBN 없으면 → 서명 대조 후 동명이서는 출판사로 구분
    """
    # 소장 목록 ISBN 세트
    lib_isbn_set: set[str] = set()
    if "ISBN13" in df_library.columns:
        lib_isbn_set = {
            clean_isbn(v) for v in df_library["ISBN13"]
            if clean_isbn(v) not in ("", "nan") and len(clean_isbn(v)) == 13
        }

    # 소장 목록 서명→출판사 딕셔너리
    title_to_publishers: dict[str, list[str]] = defaultdict(list)
    for _, row in df_library.iterrows():
        t = clean_text(row.get("서명", ""))
        p = clean_text(row.get("출판사", ""))
        if t and t != "nan":
            title_to_publishers[t].append(p)

    def check_row(row) -> tuple[str, str]:
        isbn      = clean_isbn(row.get("ISBN13", ""))
        title     = clean_text(row.get("서명", ""))
        publisher = clean_text(row.get("출판사", ""))

        if not title or title == "nan":
            return "⚠️ 확인 필요", "서명 정보 없음"

        # ── ISBN 대조 (양쪽 모두 ISBN 있을 때)
        if isbn and isbn != "nan" and len(isbn) == 13 and lib_isbn_set:
            if isbn in lib_isbn_set:
                return "❌ 소장 중(중복)", "ISBN 일치"
            return "✅ 신규(구입 가능)", "ISBN 불일치"

        # ── 서명 대조 (ISBN 없을 때 폴백)
        matched_publishers = title_to_publishers.get(title)
        if not matched_publishers:
            return "✅ 신규(구입 가능)", "서명 불일치"
        if len(matched_publishers) == 1:
            return "❌ 소장 중(중복)", "서명 일치"
        # 동명이서 → 출판사로 구분
        if publisher in matched_publishers:
            return "❌ 소장 중(중복)", "서명+출판사 일치"
        return "✅ 신규(구입 가능)", "동명이서(출판사 불일치)"

    results = df_planned.apply(check_row, axis=1, result_type="expand")
    df_planned = df_planned.copy()
    df_planned["중복여부"] = results[0]
    df_planned["대조근거"] = results[1]

    extra = [c for c in df_planned.columns if c not in ("중복여부", "대조근거")]
    return df_planned[["중복여부", "대조근거"] + extra]

# ────────────────────────────────────────────
# API 키 — Streamlit Secrets에서 로드
# ────────────────────────────────────────────
ttb_key = st.secrets.get("ALADIN_TTB_KEY", "")

with st.sidebar:
    st.header("⚙️ 설정")
    if ttb_key:
        st.success("알라딘 API 키 연결됨 ✅")
    else:
        st.warning("알라딘 API 키 없음\nISBN 자동 조회가 비활성화됩니다.")
    st.divider()
    st.markdown("**알라딘 API 키 발급**")
    st.markdown("[신청 페이지 바로가기 →](https://www.aladin.co.kr/ttb/wapi/wapiapply.aspx)")

# ────────────────────────────────────────────
# UI — 파일 업로드
# ────────────────────────────────────────────
st.divider()
col1, col2 = st.columns(2)
with col1:
    st.markdown("**① 알라딘 구입 예정 목록**")
    planned_file = st.file_uploader(
        "엑셀 파일 (.xlsx)",
        type=["xlsx", "csv"],
        key="planned",
    )
with col2:
    st.markdown("**② 도서관 현재 소장 목록**")
    library_file = st.file_uploader(
        "엑셀 또는 CSV (.xlsx / .csv)",
        type=["xlsx", "csv"],
        key="library",
    )

# ────────────────────────────────────────────
# 메인 로직
# ────────────────────────────────────────────
if planned_file and library_file:

    df_planned_raw = load_file(planned_file)
    df_library_raw = load_file(library_file)

    if df_planned_raw is None or df_library_raw is None:
        st.stop()

    st.success(
        f"✅ 구입 예정 목록 **{len(df_planned_raw):,}건** | "
        f"소장 목록 **{len(df_library_raw):,}건** 로드 완료"
    )

    # ── 컬럼 매핑 UI
    st.divider()
    st.subheader("🗂️ 컬럼 매핑 확인")
    with st.expander("컬럼 매핑 설정 열기", expanded=True):
        tab_a, tab_l = st.tabs(["알라딘 구입 예정 목록", "소장 목록"])
        with tab_a:
            df_planned = show_mapping_editor(df_planned_raw, ALADDIN_REQUIRED, "알라딘")
        with tab_l:
            df_library = show_mapping_editor(df_library_raw, LIBRARY_REQUIRED, "소장목록")

    # ── ISBN 자동 조회 (소장 목록)
    st.divider()
    st.subheader("🔍 소장 목록 ISBN 자동 조회")

    lib_has_isbn = (
        "ISBN13" in df_library.columns
        and df_library["ISBN13"].apply(
            lambda x: len(clean_isbn(x)) == 13
        ).sum() > 0
    )

    if lib_has_isbn:
        st.info("소장 목록에 이미 ISBN 정보가 있습니다. 조회를 건너뜁니다.")
    elif not ttb_key:
        st.warning(
            "알라딘 API 키가 없어 ISBN 자동 조회를 건너뜁니다. "
            "사이드바에서 키를 입력하면 대조 정확도가 높아집니다."
        )
    else:
        st.info(
            f"소장 목록 **{len(df_library)}건**에 대해 알라딘 API로 ISBN을 조회합니다. "
            f"건당 약 0.3초 소요 (예상 {len(df_library) * 0.3:.0f}초)."
        )
        if st.button("📡 ISBN 자동 조회 시작", use_container_width=True):
            df_library = enrich_library_with_isbn(df_library.copy(), ttb_key)
            st.session_state["df_library_enriched"] = df_library

            # ISBN 붙은 소장 목록 다운로드 제공
            st.download_button(
                "💾 ISBN 추가된 소장 목록 다운로드",
                data=to_excel_bytes(df_library),
                file_name="소장목록_ISBN추가.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    # 세션에 저장된 enriched 소장 목록 있으면 사용
    if "df_library_enriched" in st.session_state:
        df_library = st.session_state["df_library_enriched"]

    # ── 대조 실행
    st.divider()
    if st.button("🔎 중복 체크 실행", type="primary", use_container_width=True):
        with st.spinner("대조 중..."):
            final_df = run_duplicate_check(df_planned.copy(), df_library.copy())
        st.session_state["final_df"] = final_df

    # ── 결과 표시
    if "final_df" in st.session_state:
        final_df = st.session_state["final_df"]

        total    = len(final_df)
        new_cnt  = int((final_df["중복여부"] == "✅ 신규(구입 가능)").sum())
        dup_cnt  = int((final_df["중복여부"] == "❌ 소장 중(중복)").sum())
        warn_cnt = int((final_df["중복여부"] == "⚠️ 확인 필요").sum())

        # 대조 방식 표시
        lib_isbn_count = 0
        if "ISBN13" in df_library.columns:
            lib_isbn_count = df_library["ISBN13"].apply(
                lambda x: len(clean_isbn(x)) == 13
            ).sum()

        if lib_isbn_count > 0:
            st.info(f"📌 대조 방식: **ISBN 기준** (소장 목록 중 ISBN 확보 {lib_isbn_count}건)")
        else:
            st.info("📌 대조 방식: **서명 → 출판사 기준** (ISBN 미확보)")

        st.subheader("📊 대조 결과 요약")
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("전체",               f"{total:,}권")
        m2.metric("🟢 신규(구입 가능)", f"{new_cnt:,}권")
        m3.metric("🔴 소장 중(중복)",   f"{dup_cnt:,}권")
        m4.metric("⚠️ 확인 필요",       f"{warn_cnt:,}권")

        tab1, tab2, tab3, tab4 = st.tabs([
            f"전체 ({total:,}권)",
            f"🟢 신규 ({new_cnt:,}권)",
            f"🔴 소장 중 ({dup_cnt:,}권)",
            f"⚠️ 확인 필요 ({warn_cnt:,}권)",
        ])

        def highlight(df):
            return df.style.apply(
                lambda row: [
                    "background-color:#fff2cc;" if "소장 중" in str(row["중복여부"])
                    else "background-color:#fce4d6;" if "확인 필요" in str(row["중복여부"])
                    else "" for _ in row
                ], axis=1,
            )

        with tab1:
            st.dataframe(highlight(final_df), use_container_width=True, height=400)
        with tab2:
            d = final_df[final_df["중복여부"] == "✅ 신규(구입 가능)"]
            st.dataframe(d, use_container_width=True, height=400) if not d.empty else st.info("신규 도서가 없습니다.")
        with tab3:
            d = final_df[final_df["중복여부"] == "❌ 소장 중(중복)"]
            st.dataframe(d, use_container_width=True, height=400) if not d.empty else st.info("중복 도서가 없습니다.")
        with tab4:
            d = final_df[final_df["중복여부"] == "⚠️ 확인 필요"]
            if not d.empty:
                st.warning("서명 정보가 없어 자동 대조가 불가합니다. 수동으로 확인하세요.")
                st.dataframe(d, use_container_width=True, height=400)
            else:
                st.info("확인 필요 항목이 없습니다.")

        st.divider()
        st.subheader("💾 결과 다운로드")
        d1, d2, d3 = st.columns(3)
        with d1:
            st.download_button(
                "📥 전체 결과",
                data=to_excel_bytes(final_df),
                file_name="수서_중복체크_전체.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with d2:
            st.download_button(
                "🟢 신규 목록만",
                data=to_excel_bytes(final_df[final_df["중복여부"] == "✅ 신규(구입 가능)"]),
                file_name="수서_중복체크_신규.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with d3:
            st.download_button(
                "🔴 소장중 목록만",
                data=to_excel_bytes(final_df[final_df["중복여부"] == "❌ 소장 중(중복)"]),
                file_name="수서_중복체크_소장중.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

else:
    st.info("👆 두 파일을 모두 업로드하면 대조를 시작합니다.")
    with st.expander("📋 사용 방법"):
        st.markdown("""
1. **사이드바**에 알라딘 API 키를 입력하세요. (없어도 기본 대조는 가능)
2. **알라딘 구입 예정 목록**과 **소장 목록** 파일을 업로드하세요.
3. **컬럼 매핑**을 확인하세요. 자동 감지되지만 수정 가능합니다.
4. API 키가 있으면 **ISBN 자동 조회**로 소장 목록에 ISBN을 추가하세요.
5. **중복 체크 실행** 버튼을 누르세요.

| 대조 우선순위 | 조건 |
|-------------|------|
| ISBN 대조 | 소장 목록에 ISBN 있을 때 (가장 정확) |
| 서명 대조 | ISBN 없을 때 |
| 서명+출판사 대조 | 동명이서 2권 이상일 때 |
        """)
